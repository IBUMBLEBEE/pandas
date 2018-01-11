#! /usr/bin/env python
# -*- coding:utf-8 -*-

import string
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook


thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="1E90FF")


def write_to_execl(data):
    """
    使用numpy模块属性写入文件csv
    :param data: numpy的多维数组写入文件
    :return: None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Dell Server info"

    for row in xrange(0, len(data)):
        col = [string.capwords(letter) for letter in string.lowercase][0:len(data[row])]
        style_range(sheet, '%s%d:%s%d' % (col[0], 1, col[-1], 1), fill=fill)
        for column in col:
            sheet['%s%d' % (column, 1)].alignment = Alignment(horizontal='center', vertical='center')
        for col_num in xrange(0, len(col)):
            sheet.column_dimensions['%s' % col[col_num]].width = float(19.13)
            sheet['%s%d' % (col[col_num], row+1)] = data[row][col_num]
    wb.save(r'idc.xlsx')


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border 边框
    :param fill: An openpyxl PatternFill or GradientFill 填充和合并
    :param font: An openpyxl Font object 字体
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


if __name__ == '__main__':
    mydata = [("IDC-IP", "CPU-avg/GHz/month", "CPU-max/GHz/month", "MEM-pfree/%/last", "MEM-free/GB/last", "Disk-free/GB/last"),
              ('10.10.249.10', 4.65790046192, 16.582, 5.47822217663, 7.0071105957, 0.7575),
              ('10.10.249.11', 2.30129901367, 10.47, 4.28648433177, 5.48278045654, 28.9361),
              ('10.10.249.12', 0.971800017934, 5.876, 17.700913277, 22.6409835815, 42.6654),
              ('10.10.249.13', 1.29690945985, 18.615, 41.7125120435, 53.3538742065, 17.1958),
              ('10.10.249.14', 3.39727123927, 18.699, 2.55031976595, 3.26207733154, 28.8325),
              ('10.10.249.15', 1.04023017656, 14.262, 5.85849965324, 7.49352264404, 59.2949),
              ('10.10.249.16', 1.40527506894, 16.698, 3.4481779583, 4.41051483154, 28.6894),
              ('10.10.249.17', 1.38654620869, 11.555, 3.21760553135, 4.11559295654, 46.8353),
              ('10.10.249.18', 1.32283357618, 14.822, 8.09780074012, 10.3577804565, 53.6543),
              ('10.10.249.19', 1.08619706344, 9.36, 6.65863178384, 8.51696014404, 35.7101),
              ('10.10.249.20', 0.968379295648, 12.758, 3.46115720088, 4.42711639404, 32.7337),
              ('10.10.249.21', 1.20214113717, 8.312, 5.92263238127, 7.57555389404, 19.7813),
              ('10.10.249.22', 1.4111660126, 15.972, 4.77969554968, 6.11363983154, 37.2000),
              ('10.10.249.23', 1.86161777628, 11.407, 9.33159226974, 11.9359054565, 52.0863),
              ('10.10.249.24', 1.34119152037, 12.497, 8.58795801859, 10.9847335815, 37.3067),
              ('10.10.249.25', 1.51824339834, 12.548, 2.49840279564, 3.19567108154, 31.3535),
              ('10.10.249.26', 1.16131365032, 6.567, 5.81568270503, 7.4387512207, 37.2000),
              ('10.10.249.27', 1.17232801705, 8.369, 6.09282789012, 7.7932434082, 37.2000),
              ('10.10.249.28', 1.18459534863, 14.708, 2.6059902584, 3.3332824707, 21.4058),
              ('10.10.249.29', 1.18306256866, 6.67, 3.86116018758, 4.9387512207, 21.4062),
              ('10.10.249.30', 1.18040418759, 8.182, 7.89694379197, 10.1008605957, 37.2000),
              ('10.10.249.31', 1.00800226483, 5.449, 6.24628844715, 7.9895324707, 37.1997),
              ('10.10.249.32', 1.09502293362, 7.518, 4.605558412, 5.8908996582, 37.3067),
              ('10.10.249.33', 1.39950151328, 5.263, 6.67002282105, 8.5315246582, 37.3063),
              ('10.10.249.34', 0.996674519737, 5.539, 5.93478642094, 7.5910949707, 37.3067),
              ('10.10.249.35', 1.06712825629, 8.816, 5.99204782281, 7.6643371582, 37.3067),
              ('10.10.249.36', 1.44000269004, 15.13, 5.9042470066, 7.5520324707, 37.3063),
              ('10.10.249.37', 1.01796274127, 8.778, 7.04489413202, 9.0110168457, 37.3067),
              ('10.10.249.38', 1.01593017261, 7.677, 5.52403129813, 7.0657043457, 37.3063),
              ('10.10.249.39', 6.18788294532, 11.92, 3.26036068337, 4.17028045654, 16.5911),
              ('10.10.249.40', 1.36093574874, 17.708, 26.4451053493, 33.825553894, 49.2460),
              ('10.10.249.41', 0.941382964889, 14.153, 3.95284144907, 5.05602264404, 13.5100),
              ('10.10.249.42', 1.22652760777, 15.323, 4.77435115567, 6.10680389404, 19.4610),
              ('10.10.249.43', 2.77512438871, 16.782, 4.29793660463, 5.49742889404, 42.8132),
              ('10.10.249.44', 1.61641701603, 13.876, 17.5627225178, 22.464225769, 43.5853),
              ('10.10.249.45', 1.43482164077, 16.024, 6.18527117223, 7.91149139404, 21.8951),
              ('10.10.249.46', 2.34903281996, 15.839, 14.0147519687, 17.9260559082, 31.7422),
              ('10.10.249.47', 2.71414985093, 18.714, 7.6579728748, 9.7951965332, 38.3086),
              ('10.10.249.48', 3.97550140922, 14.18, 7.60687997678, 9.72985076904, 39.3213),
              ('10.10.249.49', 1.62244472351, 14.782, 4.27503205891, 5.46813201904, 23.1643),
              ('10.10.249.50', 0.531264941356, 14.444, 17.9506198784, 45.9368095398, 2.4444),
              ('10.10.249.51', 0.816617627939, 17.321, 2.34662584483, 6.00516891479, 2.4454),
              ('10.10.249.52', 0.787856242854, 18.127, 1.40332060881, 3.59119033813, 2.4455),
              ('10.10.249.53', 0.807874347076, 17.285, 1.40370221774, 3.59216690063, 2.4455),
              ('10.10.249.54', 2.23959194333, 16.197, 1.42046068172, 3.63505172729, 2.5521),
              ('10.10.249.55', 0.823491044408, 17.884, 1.56893888198, 4.01501846313, 2.5521),
              ('10.10.249.56', 9.73704820628, 19.855, 6.4242419326, 8.21715545654, 29.9789),
              ('10.10.249.57', 1.93258051788, 12.008, 6.63267329869, 8.48375701904, 51.6986),
              ('10.10.249.58', 1.51153504898, 16.05, 6.00966965501, 7.68688201904, 37.0905),
              ('10.10.249.59', 2.34500047079, 12.979, 5.57448328627, 7.13024139404, 37.7950),
              ('10.10.249.60', 2.65108827617, 14.027, 9.15980817682, 11.716178894, 70.3025),
              ('10.10.249.61', 2.60772487164, 14.056, 5.1057035838, 6.53063201904, 70.8188),
              ('10.10.249.62', 1.75266594929, 7.435, 5.17511848935, 6.6194152832, 54.4011),
              ('10.10.249.63', 0.971544035874, 11.179, 22.5387044235, 28.8284950256, 13.0898),
              ('10.10.249.64', 1.3429727397, 13.113, 12.3055615378, 15.7396278381, 47.4988),
              ('10.10.249.65', 0.883029061554, 10.1, 50.498707827, 64.5911903381, 74.2202),
              ('10.10.249.66', 0.805610976594, 10.126, 50.7804380161, 64.9515419006, 60.6829)
              ]
    write_to_execl(mydata)