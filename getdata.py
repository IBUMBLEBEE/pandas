#! /usr/bin/env python
# -*- coding:utf-8 -*-

import calendar
import datetime
import mysql.connector
import numpy as np
from pyzabbix import ZabbixAPI
import re
import time
import string
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl import Workbook, load_workbook

from pyVmomi import vim
from pyVim.connect import SmartConnectNoSSL, Disconnect
import random
import atexit


url = "http://ip:port"
user = "Admin"
password = "*******"

vcenter_host = 'ip'
vcenter_user = 'user'
vcenter_password = '********'
vcenter_port = 443

# 颜色设置
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="1E90FF")


class ZabbixInfoEsxi(object):
    def __init__(self):
        self.url = url
        self.user = user
        self.password = password
        self.hostidlist = []
        self.itemidslist = []
        self.hostipslist = []
        self.zapi = ZabbixAPI(self.url)
        self.zapi.login(self.user, self.password)

    def get_host_id(self):
        """
        先获取主机IP，通过ip获取hostid。通过IP的 sorted 排序
        :return: hostid is list, host ip is list
        """
        dellhost = re.compile(r'(?<![\.\d])(?:\d{1,3}\.){3}\d{1,3}(?![\.\d])')
        for h in self.zapi.host.get(output="extend"):
            for ip in dellhost.findall(h["name"]):
                if "10.10.249" in ip:
                    # print(h["name"], h["hostid"])
                    self.hostipslist.append(h["name"])
        ip = self.zapi.do_request("host.get", {"filter": {"name": sorted(self.hostipslist)}})
        for tmp in ip["result"]:
            self.hostidlist.append(tmp["hostid"])
        return self.hostidlist, sorted(self.hostipslist)

    def get_itemid_from_item(self, hostid, key):
        """

        :param hostid: list
        :param key: string
        :return: itemids is list
        """
        for h in hostid:
            object_itemids = self.zapi.do_request("item.get", {"output": "extend", "hostids": "%s" % h,
                                                               "search": {"key_": "%s" % key}})
            self.itemidslist.append(object_itemids["result"][0]["itemid"])
        return self.itemidslist


class FillExcel(object):
    def __init__(self, filename):
        """
        填充Excel表格
        :param filename: 读取文件路径，string
        """
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
        self.value_info = dict()
        self.read_excel_data()

    def read_excel_data(self):
        """
        读取生成的Excel
        :return: None
        """
        columns = [string.capwords(letter) for letter in string.lowercase[0:self.ws.max_column]]

        for column in columns:
            if column != 'A' and 'free' not in self.ws['%s1' % column].value:
                self.compete_excel_unit_used(column, [row for row in xrange(2, self.ws.max_row+1)])
            elif column != 'A' and 'free' in self.ws['%s1' % column].value:
                self.compete_excel_unit_free(column, [row for row in xrange(2, self.ws.max_row+1)])
            else:
                pass

    def compete_excel_unit_used(self, col, rows):
        """
        对已使用的参数进行颜色填充，如CPU平均值和最大值，磁盘IO的最大值
        :param col: Excel列, string
        :param rows: Excel行，list
        :return: None
        """
        for row1 in rows:
            self.value_info['%s%s' % (col, row1)] = self.ws['%s%s' % (col, row1)].value
        tmp_list = sorted(self.value_info.values(), reverse=True)[0:3]
        for key1 in self.value_info.keys():
            if self.value_info[key1] >= min(tmp_list):
                self.ws[key1].font = Font(color=colors.RED, italic=True)
            else:
                pass

    def compete_excel_unit_free(self, col, rows):
        """
        对剩余的参数进行颜色填充，如内存剩余百分比和内存剩余GB，磁盘剩余百分比
        :param col: Excel列, string
        :param rows: Excel行，list
        :return: None
        """
        for row1 in rows:
            self.value_info['%s%s' % (col, row1)] = self.ws['%s%s' % (col, row1)].value
        tmp_list = sorted(self.value_info.values(), reverse=True)[0:5]
        for key1 in self.value_info.keys():
            if self.value_info[key1] >= min(tmp_list):
                self.ws[key1].font = Font(color=colors.DARKGREEN, italic=True)
            else:
                pass

    def save_excel_file(self, save_filename):
        """
        保存文件
        :param save_filename: 保存文件路径, string
        :return: None
        """
        self.wb.save(save_filename)


def cpu_usage_history_data(hid, time_from, time_till):
    """
    Get cpu history data
    :param hid: string
    :param time_from: start timestamp
    :param time_till: end timestamp
    :return: a day data, list
    """
    cpu_usage_item = "vmware.hv.cpu.usage[{$URL},{HOST.HOST}]"
    datalist = []
    zapi = ZabbixAPI(url)
    zapi.login(user, password)
    itemid_obj = zapi.do_request("item.get", {"output": "extend", "hostids": "%s" % hid,
                                                    "search": {"key_": "%s" % cpu_usage_item}})
    itemid = itemid_obj["result"][0]["itemid"]
    cnx = mysql.connector.connect(host="10.10.255.253", user="root", password="n%KAuOa847ga", database="zabbix")
    cursor = cnx.cursor(buffered=True)
    query = ('''SELECT `value` FROM zabbix.history_uint WHERE itemid = %s AND clock >= %s AND clock <= %s''')
    cursor.execute(query, (itemid, time_from, time_till))
    for value in cursor:
        datalist.append(value[0])
    cursor.close()
    cnx.close()
    cpu_agv_data, cpu_max_data = calculation_unit(datalist)
    return cpu_agv_data, cpu_max_data


def disk_free_history_data(hid, disk_item):
    """
    Get memory last data
    :param hid: list
    :param disk_item: string
    :return: Free space on datastore (percentage)(last value), list
    """
    itemidslist = []
    zapi = ZabbixAPI(url)
    zapi.login(user, password)
    for h in hid:
        last_value = zapi.do_request("item.get", {"output": "extend", "hostids": "%s" % h,
                                                  "search": {"key_": "%s" % disk_item}})
        itemidslist.append(last_value["result"][0]["lastvalue"])
    return itemidslist


def memory_total_data(hid):
    """
    Get memory last data
    :param hid:
    :return:
    """
    memory_total = "vmware.hv.hw.memory[{$URL},{HOST.HOST}]"
    memory_usage = "vmware.hv.memory.used[{$URL},{HOST.HOST}]"
    zapi = ZabbixAPI(url)
    zapi.login(user, password)

    total_last_value = zapi.do_request("item.get", {"output": "extend", "hostids": "%s" % hid,
                                       "search": {"key_": "%s" % memory_total}})
    memory_total_last_value = total_last_value["result"][0]["lastvalue"]

    usage_last_value = zapi.do_request("item.get", {"output": "extend", "hostids": "%s" % hid,
                                                    "search": {"key_": "%s" % memory_usage}})
    memory_usage_last_value = usage_last_value["result"][0]["lastvalue"]
    # print float(memory_total_last_value)/1073741824, memory_usage_last_value
    memory_pfree_last_value = (float(memory_total_last_value)-float(memory_usage_last_value))/\
                              float(memory_total_last_value)
    memory_free_last_value = float(memory_pfree_last_value)*float(memory_total_last_value)/1073741824
    return memory_pfree_last_value*100, memory_free_last_value


# 写入 excel
def write_to_excel(data):
    """
    使用numpy模块属性写入文件csv
    :param data: numpy的多维数组写入文件
    :return: None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Dell Server info"

    for row in xrange(0, len(data)):
        col = [string.capwords(letter) for letter in string.lowercase][0:len(data[row])]
        style_range(sheet, '%s%d:%s%d' % (col[0], 1, col[-1], 1), fill=fill)
        for column in col:
            sheet['%s%d' % (column, 1)].alignment = Alignment(horizontal='center', vertical='center')
        for col_num in xrange(0, len(col)):
            sheet.column_dimensions['%s' % col[col_num]].width = float(19.13)
            sheet['%s%d' % (col[col_num], row+1)] = data[row][col_num]
    wb.save(r'idc.xlsx')


# excel 单元格处理
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border 边框
    :param fill: An openpyxl PatternFill or GradientFill 填充和合并
    :param font: An openpyxl Font object 字体
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def connect_vc_return_data(vc_host, vc_user, vc_password, vc_port, start, end):
    """
    Connect vcenter data Dell server performance data.
    :param vc_host: vcenter host. string
    :param vc_user: vcenter user. string
    :param vc_password: vcenter password. string
    :param vc_port: vcenter port. string
    :param start: The starting time for collecting data. datetime
    :param end: The data collection end time. datetime
    :return: dict, as {'172.30.10.1': '1520'}
    """
    mydata = {}

    # Connect to the host without NoSSL signing
    try:
        si = SmartConnectNoSSL(
            host=vc_host,
            user=vc_user,
            pwd=vc_password,
            port=int(vc_port))
        atexit.register(Disconnect, si)

    except IOError as e:
        pass

    if not si:
        raise SystemExit("Unable to connect to host with supplied info.")

    content = si.RetrieveContent()
    perfManager = content.perfManager

    # create a mapping from performance stats to their counterIDs
    # counterInfo: [performance stat => counterId]
    # performance stat example: cpu.usagemhz.LATEST
    # counterId example: 6
    counterInfo = {}
    for c in perfManager.perfCounter:
        prefix = c.groupInfo.key
        fullName = c.groupInfo.key + "." + c.nameInfo.key + "." + c.rollupType
        counterInfo[fullName] = c.key

    # create a list of vim.VirtualMachine objects so
    # that we can query them for statistics
    container = content.rootFolder
    # viewType = [vim.VirtualMachine]
    viewType = [vim.HostSystem]
    recursive = True

    containerView = content.viewManager.CreateContainerView(container,
                                                            viewType,
                                                            recursive)
    children = containerView.view

    # Loop through all the VMs
    for child in children:
        # Get all available metric IDs for this VM
        counterIDs = [m.counterId for m in
                      perfManager.QueryAvailablePerfMetric(entity=child)]
        # Using the IDs form a list of MetricId
        # objects for building the Query Spec
        metricIDs = [vim.PerformanceManager.MetricId(counterId=c, instance="*") for c in counterIDs]

        # Build the specification to be used
        # for querying the performance manager
        spec = vim.PerformanceManager.QuerySpec(entity=child,
                                                startTime=start,
                                                endTime=end,
                                                metricId=metricIDs
                                                # intervalId=7200,
                                                )
        # Query the performance manager
        # based on the metrics created above
        result = perfManager.QueryStats(querySpec=[spec])
        for r in result:
            for val in result[0].value:
                if counterInfo.keys()[counterInfo.values().index(val.id.counterId)] == 'disk.usage.average':
                    mydata['%s' % child.summary.config.name] = '%s' % str(val.value[0])
    return mydata


# 控制中心
def controller():
    multidimensional_array = []
    zbx = ZabbixInfoEsxi()
    hostids, hostips = zbx.get_host_id()

    # 时间处理
    time_from, time_till = handler_datetime()

    first_line = ["IDC-IP", "CPU-avg/GHz/month", "CPU-max/GHz/month", "MEM-pfree/%/last",
                  "MEM-free/GB/last", "Disk-pfree/%/last", "Disk-I/O KBps/avg"]
    multidimensional_array.append(first_line)

    # Disk Info 计算磁盘最新剩余和磁盘I/O
    start_time = datetime.datetime.fromtimestamp(float(time_from))
    end_time = datetime.datetime.fromtimestamp(float(time_till))
    disk_last_data = disk_free_history_data(hid=hostids, disk_item="pfree")
    # return data from prd vcenter
    diskio = connect_vc_return_data(vc_host=vcenter_host, vc_user=vcenter_user, vc_password=vcenter_password,
                                    vc_port=vcenter_port, start=start_time, end=end_time)
    print(diskio)

    for number in xrange(0, len(hostids)):
        # Memory
        mem_pfree_final_data, mem_free_final_data = memory_total_data(hostids[number])

        # CPU
        cpu_avg_final_data, cpu_max_final_data = cpu_usage_history_data(hid=hostids[number], time_from=time_from,
                                                                        time_till=time_till)

        print "from: {}".format(time_from), " to: {}".format(time_till), " hostip: {}".format(hostips[number]),\
              " cpu-avg: {}".format(cpu_avg_final_data), " cpu-max: {}".format(cpu_max_final_data),\
              " memory-pfree: {}%".format(mem_pfree_final_data), " memory-free: {}".format(mem_free_final_data),\
              " Disk: {}%".format(disk_last_data[number]), " I/O: {}KBps".format(diskio['{}'.format(hostips[number])])

        # 构建一维数组
        one_dimensional_array = (hostips[number], cpu_avg_final_data, cpu_max_final_data,
                                 mem_pfree_final_data, mem_free_final_data, float(disk_last_data[number]),
                                 float(diskio['{}'.format(hostips[number])]))

        # 构建多维数组
        multidimensional_array.append(one_dimensional_array)

    write_to_excel(multidimensional_array)

    # 充填Excel字体
    fill_excel = FillExcel('idc.xlsx')
    fill_excel.save_excel_file('idc%s.xlsx' % random.randrange(20, 100, 1))
    # return multidimensional_array


# 月数据计算单元，计算CPU和MEMORY
def calculation_unit(list_data):
    """
    使用numpy模块的多维数组属性进行计算
    :param list_data: list
    :return: Maximum and average value
    """
    # print(list_data)
    list_array = np.array(list_data, dtype=np.float64)
    agv_data = (list_array.sum()/list_array.size)/(1000*1000*1000)
    max_data = list_array.max()/(1000*1000*1000)
    return agv_data, max_data


# 时间处理，间隔为一个月
def handler_datetime():
    """
    获取上月每天的时间开始和结束时间戳
    :return: list
    """
    last_month = (datetime.date.today().replace(day=1) - datetime.timedelta(1)).replace(day=1)
    last_days = calendar.monthrange(last_month.year, last_month.month)[-1]
    first_day = '%s-%s-%s 00:00:00' % (last_month.year, last_month.month, last_month.day)
    last_day = '%s-%s-%s 23:59:59' % (last_month.year, last_month.month, last_days)
    start = str(time.mktime(time.strptime(first_day, '%Y-%m-%d %H:%M:%S'))).split('.')[0]
    end = str(time.mktime(time.strptime(last_day, '%Y-%m-%d %H:%M:%S'))).split('.')[0]
    return start, end


if __name__ == '__main__':
    controller()
